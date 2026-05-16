import sys
import re
import csv
import os
import re
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter

# Increase CSV field size limit to handle large fields
csv.field_size_limit(10 * 1024 * 1024)  # 10 MB, adjust as needed

#
ISAREA_REGEX = re.compile(
    r"(?i) (APS\d+|APD\d+|APC[A-Za-z_]+|SHP[A-Za-z]+|ABT\d+|ABL\d+"
    r"|ABW\d+|ABH\d+|ABF\d+|AAS\d+|AAC\d+|AAF\d+|AAD[FRD]"
    r"|AAB[OLC]|AFDEN\d+|AFMIN\d+|AFMAX\d+|AFFADE\d+"
    r"|AT\d+|AL\d+|AW\d+|AH\d+|AA\d+|ASA\d+|AS\d+|ASS\d+"
    r"|ASS\d+MS|AD[LRUD]|ASD[LRUD]|#[0-9A-Fa-f]{6,8})"
    )
#(?!nobool)|(?!notcolor)|(?!blink)[A-Za-z_]{3,19}|@[A-Za-z-]{3,19}@|
def get_colors(filepath):
    """
    Extracts color definitions from a DOF config file in the [Colors DOF] section.
    Returns a dictionary with color names as keys and their RGB values as values.
    """
    colors = {}
    with open(filepath, encoding='utf-8') as f:
        in_colors_section = False
        for line in f:
            line = line.strip()
            if line == "[Colors DOF]":
                in_colors_section = True
                continue
            if in_colors_section:
                parts = line.split('=')
                if len(parts) == 2:
                    color_name = parts[0].strip()
                    color_value = parts[1].strip()
                    colors[color_name] = color_value
    return colors


def parse_dof_config_ini(filepath):
    """
    Parses the [Config DOF] section of a directoutputconfig3?.ini file.
    Returns a list of dictionaries, one per row, with keys from the header.
    Now detects MX/RGB columns by two following empty strings after the column name in the header.
    """

    possible_colors = list(get_colors(filepath).keys())

    section = '[Config DOF]'
    in_section = False
    header = []
    mx_rgb_cols = set()
    data = []


    with open(filepath, encoding='utf-8') as f:
        for line in f:
            line = line.strip()
            if not in_section:
                if line == section:
                    in_section = True
                continue
            if line.startswith('[') and line != section:
                break  # End of section
            if not line or line.startswith(';'):
                continue  # Skip empty or comment lines
            #print(line)

            if line.startswith('#'):
                # Parse header
                header_row = next(csv.reader([line[1:]]))
                header = []
                mx_rgb_cols = set()
                i = 0
                while i < len(header_row):
                    col = header_row[i].strip()
                    if col in header:
                        col += f"_{i}"  # Make unique if duplicate
                    # Check if next two columns are empty strings
                    if i + 2 < len(header_row) and header_row[i+1].strip() == "" and header_row[i+2].strip() == "":
                        print(f"Detected MX/RGB column: {col} at index {i}")  # Debug print
                        header.append(col)
                        mx_rgb_cols.add(col)
                        i += 3  # Skip next two empty cells
                    else:
                        header.append(col)
                        i += 1
                continue
            # Parse data row
            row = next(csv.reader([line]))
            parsed_row = {}
            i = 0
            j = 0
            while i < len(header):
                #print(header[i],"   ->", j,header[i] in mx_rgb_cols,  row[j])  # Debug print to see row and header

                if header[i] in mx_rgb_cols:
                    value = row[j].strip() if j < len(row) else '0'
                    if value == "0":
                        if j + 2 < len(row):
                            parsed_row[header[i]] = value +";"+ row[j + 1].strip() +";"+ row[j + 2].strip()
                        else:
                            parsed_row[header[i]] = value +";NULL;NULL"
                        j += 3
                    elif ISAREA_REGEX.search(value) is not None or any(color.upper() in value.upper() for color in possible_colors):
                        parsed_row[header[i]] = value
                        j += 1
                    elif not (ISAREA_REGEX.search(value) is not None) and j + 2 < len(row):
                        # If the value matches the area regex, treat it as a 3 single values
                        print(row[0], ":", header[i], "   ->", header[i] in mx_rgb_cols, ":", row[j], ",",row[j + 1], ",", row[j + 2])  # Debug print
                        parsed_row[header[i]] = value +";"+ row[j + 1].strip() +";"+ row[j + 2].strip()
                        j += 3
                    else:
                        parsed_row[header[i]] = value
                        j += 1
                else:
                    value = row[j].strip() if j < len(row) else '0'
                    parsed_row[header[i]] = value
                    j += 1
                i += 1
            data.append(parsed_row)
            #if line.startswith('a7x'):
            #    break
    return header, data

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python parse_dof_config.py <config_dir>")
        sys.exit(1)
    config_dir = sys.argv[1].rstrip('/\\')
    excelFilename = os.path.basename(config_dir)
    excel_path = os.path.join(config_dir, f"{excelFilename}.xlsx")

    # Create a new workbook or load existing one
    if os.path.exists(excel_path):
        wb = load_workbook(excel_path)
    else:
        wb = Workbook()
        # Remove the default sheet if present and unused
        if "Sheet" in wb.sheetnames and len(wb.sheetnames) == 1:
            std = wb["Sheet"]
            wb.remove(std)

    for filename in os.listdir(config_dir):
        if re.match(r'directoutputconfig3.*\.ini$', filename, re.IGNORECASE):
            filepath = os.path.join(config_dir, filename)
            header, rows = parse_dof_config_ini(filepath)
            print(f"Parsed {len(rows)} rows from {filename}")
            if rows:
                sheet_name = os.path.splitext(filename)[0][:31]  # Excel sheet name max 31 chars
                # Remove sheet if it already exists
                if sheet_name in wb.sheetnames:
                    del wb[sheet_name]
                ws = wb.create_sheet(title=sheet_name)
                ws.append(header)
                for row in rows:
                    ws.append([row.get(col, "") for col in header])
                # Format as table
                end_row = len(rows) + 1  # +1 for header
                end_col = len(header)
                # Calculate Excel column letters for table range
                table_ref = f"A1:{get_column_letter(end_col)}{end_row}"
                table = Table(displayName=f"Table_{sheet_name.replace(' ', '_')}", ref=table_ref)
                style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                                       showLastColumn=False, showRowStripes=True, showColumnStripes=False)
                table.tableStyleInfo = style
                ws.add_table(table)
                print(f"Added sheet '{sheet_name}' with {len(rows)} rows.")

    wb.save(excel_path)
    print(f"Saved all configs to {excel_path}")